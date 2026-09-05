"""
OrganizadorArquivos - Ferramenta de organização inteligente de arquivos
Usa fuzzy matching para associar arquivos físicos a entradas de planilha CSV/XLSX.
"""

import sys
import os
import re
import json
import shutil
from pathlib import Path
from datetime import datetime
from copy import deepcopy

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
# O motor v2 e o loader vivem em UTILITÁRIOS/. A pasta entra no sys.path em vez
# de virar pacote: nao renomeia nada e nao depende de o nome da pasta ser um
# identificador Python valido (ele tem acento).
sys.path.insert(0, str(Path(__file__).resolve().parent / "UTILITÁRIOS"))

from matching_engine import (  # noqa: E402  (depende do sys.path acima)
    MotorMatching,
    PesosScore,
    limpar_valor_planilha,
    reescrever_sufixo,
)
from spreadsheet_loader import carregar_planilha, coluna_como_texto  # noqa: E402

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QPushButton, QLabel, QComboBox, QSlider, QFileDialog,
    QTableWidget, QTableWidgetItem, QHeaderView, QDialog, QDialogButtonBox,
    QMessageBox, QLineEdit, QCheckBox, QProgressBar, QTextEdit, QSplitter,
    QGroupBox, QScrollArea, QFrame, QAbstractItemView, QSpacerItem,
    QSizePolicy, QToolButton, QStatusBar,
)
from PySide6.QtCore import (
    Qt, QThread, Signal, QSize, QTimer,
)
from PySide6.QtGui import QColor, QFont, QIcon, QPalette


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURAÇÕES GLOBAIS
# ─────────────────────────────────────────────────────────────────────────────

CONFIG_FILE = Path(__file__).parent / "config.json"

CONFIG_PADRAO = {
    "ultimo_csv": "",
    "ultima_pasta": "",
    "threshold": 75,
    "col_matching": "",
    "col_novo_nome": "",
    "col_pasta_destino": "",
    "sufixos": [
        {"detectar": "(A)",       "reescrever": "(Ambiente)"},
        {"detectar": "(a)",       "reescrever": "(Ambiente)"},
        {"detectar": "v2",        "reescrever": "_v2"},
        {"detectar": "v3",        "reescrever": "_v3"},
        {"detectar": "face 2",    "reescrever": "_face2"},
        {"detectar": "face 3",    "reescrever": "_face3"},
        {"detectar": "face2",     "reescrever": "_face2"},
        {"detectar": "face3",     "reescrever": "_face3"},
        {"detectar": "(Ambiente)","reescrever": "(Ambiente)"},
    ],
}

# Caracteres proibidos em nomes de arquivo no Windows
CHARS_PROIBIDOS = r'\/:*?"<>|'
REGEX_CHARS_PROIBIDOS = re.compile(r'[\\/:*?"<>|]')

# ─────────────────────────────────────────────────────────────────────────────
# UTILITÁRIOS
# ─────────────────────────────────────────────────────────────────────────────

def carregar_config() -> dict:
    """Lê config.json ou retorna os valores padrão."""
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE, encoding="utf-8") as f:
                dados = json.load(f)
            # Mescla com padrão para garantir chaves novas
            cfg = deepcopy(CONFIG_PADRAO)
            cfg.update(dados)
            return cfg
        except Exception:
            pass
    return deepcopy(CONFIG_PADRAO)


def salvar_config(cfg: dict):
    """Persiste config.json ao lado do executável."""
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[AVISO] Não foi possível salvar config: {e}")


def sanitizar_nome_arquivo(nome: str) -> str:
    """Remove caracteres proibidos do Windows de um nome de arquivo."""
    return REGEX_CHARS_PROIBIDOS.sub('_', nome).strip()


def nome_sem_colisao(pasta_destino: Path, nome_arquivo: str) -> str:
    """
    Garante nome único na pasta destino adicionando _01, _02 etc se necessário.
    """
    caminho = pasta_destino / nome_arquivo
    if not caminho.exists():
        return nome_arquivo

    stem = Path(nome_arquivo).stem
    sufixo_ext = Path(nome_arquivo).suffix
    contador = 1
    while True:
        novo = f"{stem}_{contador:02d}{sufixo_ext}"
        if not (pasta_destino / novo).exists():
            return novo
        contador += 1


# ─────────────────────────────────────────────────────────────────────────────
# THREAD DE VARREDURA
# ─────────────────────────────────────────────────────────────────────────────

class ThreadVarredura(QThread):
    """Executa a varredura e matching em background para não travar a interface."""

    progresso = Signal(int)           # 0–100
    resultado_pronto = Signal(list)   # lista de dicts com correspondências
    sem_match = Signal(list)          # arquivos sem correspondência
    erro = Signal(str)

    def __init__(self, pasta_raiz: str, motor: MotorMatching,
                 df, sufixos_cfg: list[dict], threshold: int,
                 col_novo_nome: str, col_pasta_destino: str):
        super().__init__()
        self.pasta_raiz = Path(pasta_raiz)
        self.motor = motor
        # O motor v2 e puro Python e NAO carrega o DataFrame (DEC-002), entao o
        # df e a config de sufixos vem por fora — antes vinham por dentro do
        # motor antigo (motor.df, motor.sufixos_cfg).
        self.df = df
        self.sufixos_cfg = sufixos_cfg
        self.threshold = threshold
        self.col_novo_nome = col_novo_nome
        self.col_pasta_destino = col_pasta_destino

    def run(self):
        try:
            arquivos = [
                p for p in self.pasta_raiz.rglob("*")
                if p.is_file() and not p.name.startswith(".")
            ]
            total = len(arquivos)
            if total == 0:
                self.erro.emit("Nenhum arquivo encontrado na pasta selecionada.")
                return

            correspondencias = []
            sem_match = []

            for i, arq in enumerate(arquivos):
                self.progresso.emit(int((i + 1) / total * 100))

                res = self.motor.buscar(arq.name, self.threshold)

                # O v2 sempre devolve um Resultado; "sem match" e indice None.
                # metodo pode ser "código_ausente" (DEC-006): arquivo com codigo
                # claro que nao existe na planilha. Por ora cai em "Sem
                # correspondência" como qualquer outro, mas levando o motivo.
                if res.indice is None:
                    # Motivo ESTRUTURADO, nao embutido no texto: a aba precisa
                    # agrupar e contar, e parsear string de volta seria frágil.
                    sem_match.append({
                        "caminho": str(arq),
                        "motivo": res.metodo,   # "código_ausente" | "nenhum"
                        "score": res.score,
                    })
                    continue

                idx = res.indice
                linha = self.df.iloc[idx]

                # Determina novo nome (limpa * # + da planilha)
                if self.col_novo_nome and self.col_novo_nome in self.df.columns:
                    novo_nome_base = limpar_valor_planilha(str(linha[self.col_novo_nome]))
                else:
                    novo_nome_base = self.motor.nomes_planilha[idx]  # ja limpo no motor

                # Aplica mapeamento de sufixo. O v2 detecta sufixos COMPOSTOS e
                # devolve "(A) v2"; o reescrever_sufixo do motor trata cada parte
                # e preserva a ordem — o do main.py nao tratava.
                sufixo_reescrito = reescrever_sufixo(
                    res.sufixo_original, self.sufixos_cfg
                )
                if sufixo_reescrito:
                    novo_nome_base = f"{novo_nome_base} {sufixo_reescrito}"

                novo_nome = sanitizar_nome_arquivo(novo_nome_base) + arq.suffix

                # Pasta destino
                pasta_destino_val = ""
                if self.col_pasta_destino and self.col_pasta_destino in self.df.columns:
                    pasta_destino_val = str(linha[self.col_pasta_destino]).strip()

                correspondencias.append({
                    "selecionado": True,
                    "caminho_original": str(arq),
                    "nome_original": arq.name,
                    "base_detectada": res.base,
                    "sufixo_detectado": res.sufixo_original,
                    "sufixo_reescrito": sufixo_reescrito,
                    "nome_planilha": self.motor.nomes_planilha[idx],
                    "score": res.score,
                    "metodo": res.metodo,
                    "novo_nome": novo_nome,
                    "pasta_destino": pasta_destino_val,
                    "indice_df": idx,
                    # Transparencia do match — ainda NAO exibida na tabela.
                    # Guardada agora para a WO seguinte nao mexer aqui de novo.
                    "componentes": res.componentes,
                    "codigo_casado": res.codigo_casado,
                    "ambiguo": res.ambiguo,
                })

            # Ordena por score decrescente
            correspondencias.sort(key=lambda x: x["score"], reverse=True)
            self.resultado_pronto.emit(correspondencias)
            self.sem_match.emit(sem_match)

        except Exception as e:
            self.erro.emit(str(e))


# ─────────────────────────────────────────────────────────────────────────────
# THREAD DE EXECUÇÃO DAS AÇÕES
# ─────────────────────────────────────────────────────────────────────────────

class ThreadAcao(QThread):
    """Executa renomear/copiar/mover em background."""

    progresso = Signal(int)
    log_entrada = Signal(dict)
    concluido = Signal(int, int)   # (sucessos, erros)
    erro_fatal = Signal(str)

    def __init__(self, itens: list[dict], acao: str, pasta_base_destino: str):
        """
        acao: 'renomear' | 'copiar' | 'mover' | 'renomear_copiar' | 'renomear_mover'
        pasta_base_destino: pasta raiz onde subpastas serão criadas
        """
        super().__init__()
        self.itens = itens
        self.acao = acao
        self.pasta_base_destino = Path(pasta_base_destino) if pasta_base_destino else None

    def run(self):
        total = len(self.itens)
        sucessos = erros = 0

        for i, item in enumerate(self.itens):
            self.progresso.emit(int((i + 1) / total * 100))
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            caminho_orig = Path(item["caminho_original"])
            novo_nome = item["novo_nome"]
            pasta_destino_nome = item.get("pasta_destino", "")

            try:
                if not caminho_orig.exists():
                    raise FileNotFoundError(f"Arquivo não encontrado: {caminho_orig}")

                # ── Determina pasta destino ─────────────────────────────────
                # Para renomear in-place: sempre usa a pasta do arquivo original
                # Para copiar/mover: usa pasta_base_destino / subcategoria
                #                    se não há base, usa pasta_do_arquivo / subcategoria
                envolve_destino = self.acao in ("copiar", "mover", "renomear_copiar", "renomear_mover")

                if envolve_destino:
                    if self.pasta_base_destino:
                        if pasta_destino_nome:
                            pasta_final = self.pasta_base_destino / sanitizar_nome_arquivo(pasta_destino_nome)
                        else:
                            pasta_final = self.pasta_base_destino
                    else:
                        # Sem pasta base definida: usa subpasta da categoria dentro da pasta original
                        if pasta_destino_nome:
                            pasta_final = caminho_orig.parent / sanitizar_nome_arquivo(pasta_destino_nome)
                        else:
                            pasta_final = caminho_orig.parent
                else:
                    # "renomear" puro: sempre na mesma pasta
                    pasta_final = caminho_orig.parent

                pasta_final.mkdir(parents=True, exist_ok=True)

                # Nome do arquivo no destino
                nome_destino = novo_nome if "renomear" in self.acao else caminho_orig.name
                nome_final = nome_sem_colisao(pasta_final, nome_destino)
                destino = pasta_final / nome_final

                # ── Executa a ação ──────────────────────────────────────────
                if self.acao == "renomear":
                    caminho_orig.rename(caminho_orig.parent / nome_final)
                    destino_real = str(caminho_orig.parent / nome_final)

                elif self.acao == "copiar":
                    # Evita copiar sobre si mesmo
                    if destino.resolve() == caminho_orig.resolve():
                        nome_final = nome_sem_colisao(pasta_final, Path(nome_destino).stem + "_copia" + caminho_orig.suffix)
                        destino = pasta_final / nome_final
                    shutil.copy2(str(caminho_orig), str(destino))
                    destino_real = str(destino)

                elif self.acao == "mover":
                    shutil.move(str(caminho_orig), str(destino))
                    destino_real = str(destino)

                elif self.acao == "renomear_copiar":
                    # 1. Renomeia in-place
                    novo_orig = caminho_orig.parent / nome_final
                    if novo_orig != caminho_orig:
                        caminho_orig.rename(novo_orig)
                    # 2. Copia para destino (só se for pasta diferente)
                    if pasta_final.resolve() != novo_orig.parent.resolve():
                        destino_copia = pasta_final / nome_sem_colisao(pasta_final, nome_final)
                        shutil.copy2(str(novo_orig), str(destino_copia))
                        destino_real = str(destino_copia)
                    else:
                        destino_real = str(novo_orig)

                elif self.acao == "renomear_mover":
                    # 1. Renomeia in-place
                    novo_orig = caminho_orig.parent / nome_final
                    if novo_orig != caminho_orig:
                        caminho_orig.rename(novo_orig)
                    # 2. Move para destino (só se for pasta diferente)
                    if pasta_final.resolve() != novo_orig.parent.resolve():
                        destino_mov = pasta_final / nome_sem_colisao(pasta_final, nome_final)
                        shutil.move(str(novo_orig), str(destino_mov))
                        destino_real = str(destino_mov)
                    else:
                        destino_real = str(novo_orig)

                else:
                    raise ValueError(f"Ação desconhecida: {self.acao}")

                sucessos += 1
                self.log_entrada.emit({
                    "arquivo_original": str(caminho_orig),
                    "arquivo_novo": nome_final,
                    "pasta_original": str(caminho_orig.parent),
                    "pasta_nova": str(pasta_final),
                    "destino_final": destino_real,
                    "acao": self.acao,
                    "status": "sucesso",
                    "timestamp": timestamp,
                })

            except PermissionError:
                erros += 1
                self.log_entrada.emit({
                    "arquivo_original": str(caminho_orig),
                    "arquivo_novo": novo_nome,
                    "pasta_original": str(caminho_orig.parent),
                    "pasta_nova": "",
                    "destino_final": "",
                    "acao": self.acao,
                    "status": "erro: arquivo em uso",
                    "timestamp": timestamp,
                })
            except Exception as e:
                erros += 1
                self.log_entrada.emit({
                    "arquivo_original": str(caminho_orig),
                    "arquivo_novo": novo_nome,
                    "pasta_original": str(caminho_orig.parent),
                    "pasta_nova": "",
                    "destino_final": "",
                    "acao": self.acao,
                    "status": f"erro: {e}",
                    "timestamp": timestamp,
                })

        self.concluido.emit(sucessos, erros)


# ─────────────────────────────────────────────────────────────────────────────
# DIÁLOGO DE CONFIRMAÇÃO
# ─────────────────────────────────────────────────────────────────────────────

class DialogConfirmacao(QDialog):
    def __init__(self, itens: list[dict], acao: str, pasta_destino: str, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Confirmar operação")
        self.setMinimumSize(700, 420)
        layout = QVBoxLayout(self)

        labels_acao = {
            "renomear": "Renomear",
            "copiar": "Copiar para pasta",
            "mover": "Mover para pasta",
            "renomear_copiar": "Renomear + Copiar para pasta",
            "renomear_mover": "Renomear + Mover para pasta",
        }

        resumo = (
            f"<b>Ação:</b> {labels_acao.get(acao, acao)}<br>"
            f"<b>Arquivos selecionados:</b> {len(itens)}<br>"
            f"<b>Pasta base de destino:</b> {pasta_destino or '(mesma pasta do arquivo)'}"
        )
        lbl = QLabel(resumo)
        lbl.setTextFormat(Qt.RichText)
        layout.addWidget(lbl)

        tabela = QTableWidget(len(itens), 3)
        tabela.setHorizontalHeaderLabels(["Nome original", "Novo nome", "Pasta destino"])
        tabela.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        tabela.setEditTriggers(QAbstractItemView.NoEditTriggers)
        tabela.setAlternatingRowColors(True)

        for row, item in enumerate(itens):
            tabela.setItem(row, 0, QTableWidgetItem(item["nome_original"]))
            tabela.setItem(row, 1, QTableWidgetItem(item["novo_nome"]))
            tabela.setItem(row, 2, QTableWidgetItem(item.get("pasta_destino", "")))

        layout.addWidget(tabela)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.button(QDialogButtonBox.Ok).setText("✔  Executar")
        btns.button(QDialogButtonBox.Cancel).setText("Cancelar")
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)


# ─────────────────────────────────────────────────────────────────────────────
# ABA: CORRESPONDÊNCIAS
# ─────────────────────────────────────────────────────────────────────────────

COLUNAS_TABELA = [
    "✔", "Nome original", "Base detectada", "Sufixo", "Correspondência planilha",
    "Score", "Método", "Por quê", "Novo nome (editável)", "Pasta destino",
]
IDX_CHECK = 0
IDX_NOME_ORIG = 1
IDX_BASE = 2
IDX_SUFIXO = 3
IDX_MATCH = 4
IDX_SCORE = 5
IDX_METODO = 6
IDX_PORQUE = 7
IDX_NOVO_NOME = 8
IDX_PASTA_DEST = 9


def resumir_match(item: dict) -> tuple[str, str]:
    """
    Traduz a transparência do motor em (texto curto da célula, tooltip detalhado).

    Uma coluna só, e não uma por componente: `componentes` traz cinco campos que
    ficam TODOS vazios num match por código — e match por código é a maioria dos
    casos. Cinco colunas vazias na maior parte das linhas custam legibilidade e
    não pagam nada. O detalhe completo vai no tooltip, que não ocupa espaço.
    """
    metodo = item.get("metodo", "")
    comp = item.get("componentes") or {}
    ambiguo = bool(item.get("ambiguo"))

    if metodo == "código":
        curto = f"código {item.get('codigo_casado', '')}".strip()
        detalhe = (
            f"Casou pelo código de referência: {item.get('codigo_casado', '—')}\n"
            f"Match por código é exato ou bidirecional (FIX-001) e vale 100 —\n"
            f"os componentes do score fuzzy não são calculados neste caminho."
        )
    elif metodo == "fuzzy" and comp:
        curto = (
            f"sort {comp.get('token_sort', '?')} · "
            f"wratio {comp.get('wratio', '?')} · "
            f"cob {comp.get('cobertura', '?')} · "
            f"medida {comp.get('medida', '?')}"
        )
        detalhe = (
            f"Sem código utilizável — casou por score ponderado (DEC-002).\n"
            f"  token_sort    {comp.get('token_sort', '?')}\n"
            f"  WRatio        {comp.get('wratio', '?')}\n"
            f"  cobertura     {comp.get('cobertura', '?')}%  (tokens da planilha presentes no arquivo)\n"
            f"  medida        {comp.get('medida', '?')}  (ajuste {comp.get('ajuste_medida', 0)})\n"
            f"Medida divergente penaliza forte — ver DEC-005."
        )
    else:
        curto = comp.get("motivo", "—")
        detalhe = curto

    if ambiguo:
        curto = f"⚠ ambíguo · {curto}"
        detalhe += (
            "\n\nAMBÍGUO: o segundo melhor candidato ficou dentro da margem.\n"
            "A linha continua marcada, mas confira antes de executar."
        )
    return curto, detalhe


class AbaCorrespondencias(QWidget):
    solicitar_acao = Signal(str, str)  # (acao, pasta_base_destino)

    def __init__(self, cfg: dict):
        super().__init__()
        self.cfg = cfg
        self.correspondencias: list[dict] = []
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(8)

        # ── Barra de controles ────────────────────────────────────────────
        barra = QHBoxLayout()

        self.btn_selecionar_todos = QPushButton("Selecionar todos")
        self.btn_selecionar_todos.clicked.connect(self._selecionar_todos)
        barra.addWidget(self.btn_selecionar_todos)

        self.btn_desselecionar = QPushButton("Desmarcar todos")
        self.btn_desselecionar.clicked.connect(self._desmarcar_todos)
        barra.addWidget(self.btn_desselecionar)

        barra.addSpacerItem(QSpacerItem(20, 0, QSizePolicy.Expanding))

        lbl_pasta = QLabel("Pasta base destino:")
        barra.addWidget(lbl_pasta)

        self.campo_pasta_destino = QLineEdit()
        self.campo_pasta_destino.setPlaceholderText(
            "Deixe vazio para usar pasta do próprio arquivo"
        )
        self.campo_pasta_destino.setMinimumWidth(300)
        barra.addWidget(self.campo_pasta_destino)

        btn_browse_pasta = QToolButton()
        btn_browse_pasta.setText("…")
        btn_browse_pasta.clicked.connect(self._selecionar_pasta_destino)
        barra.addWidget(btn_browse_pasta)

        layout.addLayout(barra)

        # ── Tabela de resultados ──────────────────────────────────────────
        self.tabela = QTableWidget(0, len(COLUNAS_TABELA))
        self.tabela.setHorizontalHeaderLabels(COLUNAS_TABELA)
        self.tabela.setAlternatingRowColors(True)
        self.tabela.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tabela.setSortingEnabled(True)
        self.tabela.verticalHeader().setDefaultSectionSize(26)

        hh = self.tabela.horizontalHeader()
        hh.setSectionResizeMode(IDX_CHECK, QHeaderView.Fixed)
        self.tabela.setColumnWidth(IDX_CHECK, 30)
        hh.setSectionResizeMode(IDX_NOME_ORIG, QHeaderView.Interactive)
        self.tabela.setColumnWidth(IDX_NOME_ORIG, 220)
        hh.setSectionResizeMode(IDX_BASE, QHeaderView.Interactive)
        self.tabela.setColumnWidth(IDX_BASE, 180)
        hh.setSectionResizeMode(IDX_SUFIXO, QHeaderView.Fixed)
        self.tabela.setColumnWidth(IDX_SUFIXO, 80)
        hh.setSectionResizeMode(IDX_MATCH, QHeaderView.Interactive)
        self.tabela.setColumnWidth(IDX_MATCH, 240)
        hh.setSectionResizeMode(IDX_SCORE, QHeaderView.Fixed)
        self.tabela.setColumnWidth(IDX_SCORE, 55)
        hh.setSectionResizeMode(IDX_METODO, QHeaderView.Fixed)
        self.tabela.setColumnWidth(IDX_METODO, 65)
        hh.setSectionResizeMode(IDX_NOVO_NOME, QHeaderView.Interactive)
        self.tabela.setColumnWidth(IDX_NOVO_NOME, 260)
        hh.setSectionResizeMode(IDX_PASTA_DEST, QHeaderView.Stretch)

        layout.addWidget(self.tabela)

        # ── Botões de ação ────────────────────────────────────────────────
        barra_acoes = QHBoxLayout()

        acoes = [
            ("Renomear", "renomear"),
            ("Copiar para pasta", "copiar"),
            ("Mover para pasta", "mover"),
            ("Renomear + Copiar", "renomear_copiar"),
            ("Renomear + Mover", "renomear_mover"),
        ]
        for label, acao in acoes:
            btn = QPushButton(label)
            btn.clicked.connect(lambda checked, a=acao: self._disparar_acao(a))
            barra_acoes.addWidget(btn)

        layout.addLayout(barra_acoes)

        # ── Barra de progresso ────────────────────────────────────────────
        self.progresso = QProgressBar()
        self.progresso.setVisible(False)
        layout.addWidget(self.progresso)

    def _selecionar_pasta_destino(self):
        pasta = QFileDialog.getExistingDirectory(self, "Selecionar pasta base destino")
        if pasta:
            self.campo_pasta_destino.setText(pasta)

    def _selecionar_todos(self):
        for row in range(self.tabela.rowCount()):
            item = self.tabela.item(row, IDX_CHECK)
            if item:
                item.setCheckState(Qt.Checked)
                self.correspondencias[row]["selecionado"] = True

    def _desmarcar_todos(self):
        for row in range(self.tabela.rowCount()):
            item = self.tabela.item(row, IDX_CHECK)
            if item:
                item.setCheckState(Qt.Unchecked)
                self.correspondencias[row]["selecionado"] = False

    def _disparar_acao(self, acao: str):
        pasta_destino = self.campo_pasta_destino.text().strip()
        self.solicitar_acao.emit(acao, pasta_destino)

    def popular_tabela(self, correspondencias: list[dict]):
        self.correspondencias = correspondencias
        self.tabela.setSortingEnabled(False)
        self.tabela.setRowCount(0)
        self.tabela.setRowCount(len(correspondencias))

        for row, item in enumerate(correspondencias):
            # Checkbox
            check_item = QTableWidgetItem()
            check_item.setCheckState(Qt.Checked if item["selecionado"] else Qt.Unchecked)
            check_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            self.tabela.setItem(row, IDX_CHECK, check_item)

            def _cell(texto):
                c = QTableWidgetItem(str(texto))
                c.setFlags(c.flags() & ~Qt.ItemIsEditable)
                return c

            self.tabela.setItem(row, IDX_NOME_ORIG, _cell(item["nome_original"]))
            self.tabela.setItem(row, IDX_BASE, _cell(item["base_detectada"]))
            self.tabela.setItem(row, IDX_SUFIXO, _cell(item["sufixo_detectado"]))
            self.tabela.setItem(row, IDX_MATCH, _cell(item["nome_planilha"]))

            # Score com cor
            score_item = QTableWidgetItem(str(item["score"]))
            score_item.setFlags(score_item.flags() & ~Qt.ItemIsEditable)
            score = item["score"]
            if score >= 85:
                score_item.setBackground(QColor("#c8f7c5"))
                score_item.setForeground(QColor("#1a5c14"))
            elif score >= 65:
                score_item.setBackground(QColor("#fff3cd"))
                score_item.setForeground(QColor("#7a5c00"))
            else:
                score_item.setBackground(QColor("#fad4d4"))
                score_item.setForeground(QColor("#8b1c1c"))
            # Ambiguidade e ALERTA, nao nota: mesmo com score alto a celula vai
            # de ambar, para o olho parar nela. O checkbox NAO e desmarcado —
            # mudar selecao por conta propria esconderia arquivo do usuario.
            if item.get("ambiguo"):
                score_item.setBackground(QColor("#ffe0b2"))
                score_item.setForeground(QColor("#8a4b00"))
            self.tabela.setItem(row, IDX_SCORE, score_item)

            self.tabela.setItem(row, IDX_METODO, _cell(item["metodo"]))

            # Por que esta linha casou — resumo na celula, detalhe no tooltip.
            curto, detalhe = resumir_match(item)
            porque_item = _cell(curto)
            porque_item.setToolTip(detalhe)
            score_item.setToolTip(detalhe)
            if item.get("ambiguo"):
                porque_item.setForeground(QColor("#8a4b00"))
            self.tabela.setItem(row, IDX_PORQUE, porque_item)

            # Novo nome: editável
            novo_nome_item = QTableWidgetItem(item["novo_nome"])
            self.tabela.setItem(row, IDX_NOVO_NOME, novo_nome_item)

            self.tabela.setItem(row, IDX_PASTA_DEST, _cell(item["pasta_destino"]))

        self.tabela.setSortingEnabled(True)

        # Conecta edição do novo nome de volta ao dict
        self.tabela.itemChanged.connect(self._on_item_changed)

    def _on_item_changed(self, item: QTableWidgetItem):
        row = item.row()
        col = item.column()
        if col == IDX_NOVO_NOME and row < len(self.correspondencias):
            self.correspondencias[row]["novo_nome"] = item.text()
        elif col == IDX_CHECK and row < len(self.correspondencias):
            self.correspondencias[row]["selecionado"] = (
                item.checkState() == Qt.Checked
            )

    def obter_selecionados(self) -> list[dict]:
        # Sincroniza estado atual da tabela antes de retornar
        for row in range(self.tabela.rowCount()):
            check_item = self.tabela.item(row, IDX_CHECK)
            novo_nome_item = self.tabela.item(row, IDX_NOVO_NOME)
            if row < len(self.correspondencias):
                if check_item:
                    self.correspondencias[row]["selecionado"] = (
                        check_item.checkState() == Qt.Checked
                    )
                if novo_nome_item:
                    self.correspondencias[row]["novo_nome"] = novo_nome_item.text()
        return [c for c in self.correspondencias if c.get("selecionado")]

    def mostrar_progresso(self, valor: int):
        self.progresso.setVisible(True)
        self.progresso.setValue(valor)
        if valor >= 100:
            QTimer.singleShot(1500, lambda: self.progresso.setVisible(False))


# ─────────────────────────────────────────────────────────────────────────────
# ABA: SEM CORRESPONDÊNCIA
# ─────────────────────────────────────────────────────────────────────────────

class AbaSemMatch(QWidget):
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout(self)
        lbl = QLabel("Arquivos sem correspondência encontrada na planilha:")
        layout.addWidget(lbl)
        self.lista = QTextEdit()
        self.lista.setReadOnly(True)
        layout.addWidget(self.lista)

    def popular(self, arquivos: list[dict]):
        """
        Agrupa por MOTIVO, porque os dois grupos pedem ações diferentes:
        "código não cadastrado" é trabalho na planilha (cadastrar o produto);
        "nenhuma correspondência" é trabalho no matching (baixar o threshold,
        conferir o nome). Misturar os dois numa lista só esconde essa diferença.
        Não vira aba separada: continuam sendo o mesmo destino do usuário.
        """
        if not arquivos:
            self.lista.setPlainText("(Nenhum arquivo sem correspondência — ótimo!)")
            return

        ausentes = [a for a in arquivos if a.get("motivo") == "código_ausente"]
        outros = [a for a in arquivos if a.get("motivo") != "código_ausente"]

        blocos = []
        if ausentes:
            blocos.append(
                f"CÓDIGO NÃO CADASTRADO ({len(ausentes)})\n"
                "O arquivo tem um código de referência claro que não existe na "
                "planilha.\nO motor NÃO tenta adivinhar por semelhança (DEC-006) — "
                "provavelmente\nfalta cadastrar o produto.\n\n"
                + "\n".join(a["caminho"] for a in ausentes)
            )
        if outros:
            blocos.append(
                f"SEM CORRESPONDÊNCIA ({len(outros)})\n"
                "Nenhuma linha passou do threshold. Confira o nome do arquivo ou "
                "baixe o threshold.\n\n"
                + "\n".join(
                    f"{a['caminho']}   (melhor score: {a.get('score', 0)})"
                    for a in outros
                )
            )
        self.lista.setPlainText(("\n\n" + "─" * 70 + "\n\n").join(blocos))


# ─────────────────────────────────────────────────────────────────────────────
# ABA: LOG
# ─────────────────────────────────────────────────────────────────────────────

class AbaLog(QWidget):
    def __init__(self):
        super().__init__()
        self.entradas: list[dict] = []
        layout = QVBoxLayout(self)

        barra = QHBoxLayout()
        btn_exportar = QPushButton("Exportar log como Excel")
        btn_exportar.clicked.connect(self._exportar)
        barra.addWidget(btn_exportar)

        btn_desfazer = QPushButton("↩  Desfazer última operação em massa")
        btn_desfazer.clicked.connect(self._desfazer)
        barra.addWidget(btn_desfazer)

        btn_limpar = QPushButton("Limpar log")
        btn_limpar.clicked.connect(self._limpar)
        barra.addWidget(btn_limpar)

        barra.addStretch()
        layout.addLayout(barra)

        self.tabela = QTableWidget(0, 7)
        self.tabela.setHorizontalHeaderLabels([
            "Arquivo original", "Novo nome", "Pasta original",
            "Pasta destino", "Ação", "Status", "Data/Hora",
        ])
        self.tabela.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.tabela.horizontalHeader().setStretchLastSection(True)
        self.tabela.setAlternatingRowColors(True)
        self.tabela.setEditTriggers(QAbstractItemView.NoEditTriggers)
        layout.addWidget(self.tabela)

        self._ultimo_lote_inicio: int = 0  # índice do início do último lote

    def adicionar_entrada(self, entrada: dict):
        self.entradas.append(entrada)
        row = self.tabela.rowCount()
        self.tabela.insertRow(row)
        valores = [
            entrada.get("arquivo_original", ""),
            entrada.get("arquivo_novo", ""),
            entrada.get("pasta_original", ""),
            entrada.get("pasta_nova", ""),
            entrada.get("acao", ""),
            entrada.get("status", ""),
            entrada.get("timestamp", ""),
        ]
        status = entrada.get("status", "").lower()
        if "sucesso" in status:
            cor = QColor("#c8f7c5")
        elif "erro" in status:
            cor = QColor("#fad4d4")
        else:
            cor = QColor("#fff3cd")

        for col, val in enumerate(valores):
            item = QTableWidgetItem(str(val))
            item.setBackground(cor)
            self.tabela.setItem(row, col, item)

        self.tabela.scrollToBottom()

    def marcar_inicio_lote(self):
        self._ultimo_lote_inicio = len(self.entradas)

    def _desfazer(self):
        lote = self.entradas[self._ultimo_lote_inicio:]
        if not lote:
            QMessageBox.information(self, "Desfazer", "Nenhum lote recente para desfazer.")
            return

        resp = QMessageBox.question(
            self, "Desfazer",
            f"Reverter {len(lote)} operação(ões) do último lote?\n\n"
            "Apenas operações de renomear e mover podem ser desfeitas.",
            QMessageBox.Yes | QMessageBox.No,
        )
        if resp != QMessageBox.Yes:
            return

        erros = []
        for entrada in reversed(lote):
            acao = entrada.get("acao", "")
            destino_final = entrada.get("destino_final", "")
            original = entrada.get("arquivo_original", "")
            if not destino_final or not original:
                continue
            try:
                p_destino = Path(destino_final)
                p_original = Path(original)
                if p_destino.exists() and not p_original.exists():
                    if "mover" in acao:
                        shutil.move(str(p_destino), str(p_original))
                    else:
                        p_destino.rename(p_original)
            except Exception as e:
                erros.append(str(e))

        if erros:
            QMessageBox.warning(self, "Desfazer parcial",
                                "Alguns arquivos não puderam ser revertidos:\n" + "\n".join(erros[:10]))
        else:
            QMessageBox.information(self, "Desfazer", "Operações revertidas com sucesso.")

    def _exportar(self):
        if not self.entradas:
            QMessageBox.information(self, "Log vazio", "Nenhuma operação registrada.")
            return
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        caminho, _ = QFileDialog.getSaveFileName(
            self, "Salvar log",
            f"log_organizador_{ts}.xlsx", "Excel (*.xlsx)"
        )
        if not caminho:
            return
        exportar_log_excel(self.entradas, caminho)
        QMessageBox.information(self, "Exportado", f"Log salvo em:\n{caminho}")

    def _limpar(self):
        resp = QMessageBox.question(self, "Limpar", "Limpar todo o histórico de log?",
                                    QMessageBox.Yes | QMessageBox.No)
        if resp == QMessageBox.Yes:
            self.entradas.clear()
            self.tabela.setRowCount(0)
            self._ultimo_lote_inicio = 0


def exportar_log_excel(log: list[dict], caminho_saida: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Log de Operações"
    colunas = ["Arquivo Original", "Novo Nome", "Pasta Original",
               "Pasta Destino", "Ação", "Status", "Data/Hora"]
    header_fill = PatternFill("solid", fgColor="1D4E89")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, titulo in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    cor_ok = PatternFill("solid", fgColor="D6F5D6")
    cor_err = PatternFill("solid", fgColor="FAD4D4")
    cor_skip = PatternFill("solid", fgColor="FFF3CD")

    for row_idx, e in enumerate(log, 2):
        vals = [
            e.get("arquivo_original", ""), e.get("arquivo_novo", ""),
            e.get("pasta_original", ""), e.get("pasta_nova", ""),
            e.get("acao", ""), e.get("status", ""), e.get("timestamp", ""),
        ]
        status = e.get("status", "").lower()
        fill = cor_ok if "sucesso" in status else (cor_err if "erro" in status else cor_skip)
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    wb.save(caminho_saida)


# ─────────────────────────────────────────────────────────────────────────────
# ABA: CONFIGURAÇÕES
# ─────────────────────────────────────────────────────────────────────────────

class AbaConfiguracoes(QWidget):
    config_alterada = Signal(dict)

    def __init__(self, cfg: dict):
        super().__init__()
        self.cfg = cfg
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        # Grupo: Sufixos
        grp = QGroupBox("Mapeamento de sufixos (detectar → reescrever)")
        grp_layout = QVBoxLayout(grp)

        info = QLabel(
            "Sufixos detectados no fim do nome do arquivo são removidos antes do matching\n"
            "e recolados (reescritos) depois do novo nome.\n"
            "Ex: '(A)' → '(Ambiente)'  faz  'PISO XYZ (A)'  virar  'PISO XYZ DESCRICAO (Ambiente)'"
        )
        info.setWordWrap(True)
        grp_layout.addWidget(info)

        self.tabela_sufixos = QTableWidget(0, 2)
        self.tabela_sufixos.setHorizontalHeaderLabels(["Detectar (no arquivo)", "Reescrever como"])
        self.tabela_sufixos.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tabela_sufixos.setMinimumHeight(200)
        grp_layout.addWidget(self.tabela_sufixos)

        barra = QHBoxLayout()
        btn_add = QPushButton("+ Adicionar linha")
        btn_add.clicked.connect(self._add_linha)
        barra.addWidget(btn_add)
        btn_rem = QPushButton("– Remover linha selecionada")
        btn_rem.clicked.connect(self._rem_linha)
        barra.addWidget(btn_rem)
        barra.addStretch()
        btn_salvar = QPushButton("Salvar configurações")
        btn_salvar.clicked.connect(self._salvar)
        barra.addWidget(btn_salvar)
        grp_layout.addLayout(barra)

        layout.addWidget(grp)
        layout.addStretch()

        # Preenche tabela com config atual
        self._preencher()

    def _preencher(self):
        self.tabela_sufixos.setRowCount(0)
        for entrada in self.cfg.get("sufixos", []):
            row = self.tabela_sufixos.rowCount()
            self.tabela_sufixos.insertRow(row)
            self.tabela_sufixos.setItem(row, 0, QTableWidgetItem(entrada.get("detectar", "")))
            self.tabela_sufixos.setItem(row, 1, QTableWidgetItem(entrada.get("reescrever", "")))

    def _add_linha(self):
        row = self.tabela_sufixos.rowCount()
        self.tabela_sufixos.insertRow(row)

    def _rem_linha(self):
        row = self.tabela_sufixos.currentRow()
        if row >= 0:
            self.tabela_sufixos.removeRow(row)

    def _salvar(self):
        sufixos = []
        for row in range(self.tabela_sufixos.rowCount()):
            d = self.tabela_sufixos.item(row, 0)
            r = self.tabela_sufixos.item(row, 1)
            det = d.text().strip() if d else ""
            ree = r.text().strip() if r else ""
            if det:
                sufixos.append({"detectar": det, "reescrever": ree})
        self.cfg["sufixos"] = sufixos
        salvar_config(self.cfg)
        self.config_alterada.emit(self.cfg)
        QMessageBox.information(self, "Salvo", "Configurações salvas com sucesso.")


# ─────────────────────────────────────────────────────────────────────────────
# JANELA PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

class JanelaPrincipal(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Organizador de Arquivos Inteligente")
        self.setMinimumSize(1100, 720)

        self.cfg = carregar_config()
        self.df: pd.DataFrame | None = None
        self.motor: MotorMatching | None = None
        self.thread_varredura: ThreadVarredura | None = None
        self.thread_acao: ThreadAcao | None = None

        self._build_ui()
        self._restaurar_estado()

    # ── Construção da interface ───────────────────────────────────────────

    def _build_ui(self):
        widget_central = QWidget()
        self.setCentralWidget(widget_central)
        layout_principal = QVBoxLayout(widget_central)
        layout_principal.setSpacing(8)
        layout_principal.setContentsMargins(10, 10, 10, 6)

        # ── Barra superior ────────────────────────────────────────────────
        barra_sup = QFrame()
        barra_sup.setFrameShape(QFrame.StyledPanel)
        barra_layout = QVBoxLayout(barra_sup)
        barra_layout.setSpacing(6)

        # Linha 1: arquivo + pasta
        linha1 = QHBoxLayout()

        btn_csv = QPushButton("📂  Carregar CSV / XLSX")
        btn_csv.setFixedHeight(32)
        btn_csv.clicked.connect(self._carregar_planilha)
        linha1.addWidget(btn_csv)

        self.lbl_csv = QLabel("Nenhuma planilha carregada")
        self.lbl_csv.setStyleSheet("color: gray; font-size: 12px;")
        linha1.addWidget(self.lbl_csv)

        linha1.addSpacerItem(QSpacerItem(20, 0, QSizePolicy.Expanding))

        btn_pasta = QPushButton("📁  Selecionar pasta raiz")
        btn_pasta.setFixedHeight(32)
        btn_pasta.clicked.connect(self._selecionar_pasta_raiz)
        linha1.addWidget(btn_pasta)

        self.lbl_pasta = QLabel("Nenhuma pasta selecionada")
        self.lbl_pasta.setStyleSheet("color: gray; font-size: 12px;")
        linha1.addWidget(self.lbl_pasta)

        barra_layout.addLayout(linha1)

        # Linha 2: colunas (em GroupBox com duas sublinha para não comprimir)
        grp_colunas = QGroupBox("Mapeamento de colunas da planilha")
        grp_colunas_layout = QVBoxLayout(grp_colunas)
        grp_colunas_layout.setSpacing(4)

        # Sub-linha A: matching + novo nome
        sub_a = QHBoxLayout()
        lbl_match = QLabel("Col. matching (comparação):")
        lbl_match.setFixedWidth(190)
        sub_a.addWidget(lbl_match)
        self.cb_col_matching = QComboBox()
        self.cb_col_matching.setMinimumWidth(220)
        self.cb_col_matching.setToolTip("Coluna da planilha comparada com nomes de arquivos (ex: 'Atual: ...')")
        sub_a.addWidget(self.cb_col_matching)

        sub_a.addSpacerItem(QSpacerItem(16, 0))

        lbl_nome = QLabel("Col. novo nome:")
        lbl_nome.setFixedWidth(110)
        sub_a.addWidget(lbl_nome)
        self.cb_col_novo_nome = QComboBox()
        self.cb_col_novo_nome.setMinimumWidth(220)
        self.cb_col_novo_nome.setToolTip("Coluna usada como novo nome do arquivo (deixe vazio para usar a de matching)")
        sub_a.addWidget(self.cb_col_novo_nome)
        sub_a.addStretch()
        grp_colunas_layout.addLayout(sub_a)

        # Sub-linha B: pasta destino + threshold + escanear
        sub_b = QHBoxLayout()
        lbl_dest = QLabel("Col. pasta destino:")
        lbl_dest.setFixedWidth(190)
        sub_b.addWidget(lbl_dest)
        self.cb_col_pasta_destino = QComboBox()
        self.cb_col_pasta_destino.setMinimumWidth(220)
        self.cb_col_pasta_destino.setToolTip("Coluna usada para criar subpastas (ex: Tipo, Material)")
        sub_b.addWidget(self.cb_col_pasta_destino)

        sub_b.addSpacerItem(QSpacerItem(16, 0))

        lbl_thr = QLabel("Threshold fuzzy:")
        lbl_thr.setFixedWidth(110)
        sub_b.addWidget(lbl_thr)
        self.slider_threshold = QSlider(Qt.Horizontal)
        self.slider_threshold.setRange(50, 100)
        self.slider_threshold.setValue(self.cfg.get("threshold", 75))
        self.slider_threshold.setFixedWidth(130)
        self.slider_threshold.valueChanged.connect(self._atualizar_lbl_threshold)
        sub_b.addWidget(self.slider_threshold)
        self.lbl_threshold = QLabel(f"{self.cfg.get('threshold', 75)}%")
        self.lbl_threshold.setFixedWidth(36)
        sub_b.addWidget(self.lbl_threshold)

        sub_b.addSpacerItem(QSpacerItem(8, 0))
        btn_escanear = QPushButton("🔍  Escanear e simular")
        btn_escanear.setFixedHeight(30)
        btn_escanear.clicked.connect(self._escanear)
        sub_b.addWidget(btn_escanear)
        sub_b.addStretch()
        grp_colunas_layout.addLayout(sub_b)

        barra_layout.addWidget(grp_colunas)
        layout_principal.addWidget(barra_sup)

        # ── Barra de progresso de varredura ───────────────────────────────
        self.barra_varredura = QProgressBar()
        self.barra_varredura.setVisible(False)
        self.barra_varredura.setFixedHeight(14)
        layout_principal.addWidget(self.barra_varredura)

        # ── Abas ──────────────────────────────────────────────────────────
        self.abas = QTabWidget()

        self.aba_correspondencias = AbaCorrespondencias(self.cfg)
        self.aba_correspondencias.solicitar_acao.connect(self._executar_acao)
        self.abas.addTab(self.aba_correspondencias, "Correspondências")

        self.aba_sem_match = AbaSemMatch()
        self.abas.addTab(self.aba_sem_match, "Sem correspondência")

        self.aba_configuracoes = AbaConfiguracoes(self.cfg)
        self.aba_configuracoes.config_alterada.connect(self._on_config_alterada)
        self.abas.addTab(self.aba_configuracoes, "Configurações")

        self.aba_log = AbaLog()
        self.abas.addTab(self.aba_log, "Log")

        layout_principal.addWidget(self.abas)

        # ── Status bar ────────────────────────────────────────────────────
        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self.status.showMessage("Pronto. Carregue uma planilha e selecione uma pasta para começar.")

    # ── Restaurar estado anterior ─────────────────────────────────────────

    def _restaurar_estado(self):
        self.slider_threshold.setValue(self.cfg.get("threshold", 75))
        ultimo_csv = self.cfg.get("ultimo_csv", "")
        if ultimo_csv and Path(ultimo_csv).exists():
            self._carregar_planilha(caminho=ultimo_csv)
        ultima_pasta = self.cfg.get("ultima_pasta", "")
        if ultima_pasta and Path(ultima_pasta).exists():
            self.lbl_pasta.setText(ultima_pasta)
            self.cfg["ultima_pasta"] = ultima_pasta

    # ── Slots ─────────────────────────────────────────────────────────────

    def _atualizar_lbl_threshold(self, val: int):
        self.lbl_threshold.setText(f"{val}%")
        self.cfg["threshold"] = val

    def _carregar_planilha(self, caminho: str = None):
        if not caminho:
            caminho, _ = QFileDialog.getOpenFileName(
                self, "Abrir planilha",
                self.cfg.get("ultimo_csv", ""),
                "Planilhas (*.csv *.xlsx *.xls)"
            )
        if not caminho:
            return

        try:
            # Carga SEMPRE como texto — ver FIX-005. A limpeza dos nomes de
            # coluna acontece dentro de carregar_planilha().
            self.df = carregar_planilha(caminho)
            self._preencher_comboboxes()
            self.lbl_csv.setText(f"{Path(caminho).name}  ({len(self.df)} linhas)")
            self.lbl_csv.setStyleSheet("color: green; font-size: 12px;")
            self.cfg["ultimo_csv"] = caminho
            salvar_config(self.cfg)
            self.status.showMessage(f"Planilha carregada: {len(self.df)} registros, {len(self.df.columns)} colunas.")

        except Exception as e:
            QMessageBox.critical(self, "Erro ao carregar planilha", str(e))

    def _preencher_comboboxes(self):
        if self.df is None:
            return
        colunas = ["(nenhuma)"] + list(self.df.columns)

        for cb in (self.cb_col_matching, self.cb_col_novo_nome, self.cb_col_pasta_destino):
            cb.blockSignals(True)
            cb.clear()
            cb.addItems(colunas)
            cb.blockSignals(False)

        # Tenta restaurar seleção salva
        def _set(cb, salvo):
            if salvo and salvo in self.df.columns:
                cb.setCurrentText(salvo)

        _set(self.cb_col_matching, self.cfg.get("col_matching", ""))
        _set(self.cb_col_novo_nome, self.cfg.get("col_novo_nome", ""))
        _set(self.cb_col_pasta_destino, self.cfg.get("col_pasta_destino", ""))

    def _selecionar_pasta_raiz(self):
        pasta = QFileDialog.getExistingDirectory(
            self, "Selecionar pasta raiz",
            self.cfg.get("ultima_pasta", "")
        )
        if pasta:
            self.lbl_pasta.setText(pasta)
            self.lbl_pasta.setStyleSheet("color: green; font-size: 12px;")
            self.cfg["ultima_pasta"] = pasta
            salvar_config(self.cfg)

    def _escanear(self):
        if self.df is None:
            QMessageBox.warning(self, "Atenção", "Carregue uma planilha antes de escanear.")
            return

        pasta_raiz = self.cfg.get("ultima_pasta", "")
        if not pasta_raiz or not Path(pasta_raiz).exists():
            QMessageBox.warning(self, "Atenção", "Selecione uma pasta raiz válida.")
            return

        col_matching = self.cb_col_matching.currentText()
        if col_matching == "(nenhuma)" or col_matching not in self.df.columns:
            QMessageBox.warning(self, "Atenção", "Selecione a coluna de matching.")
            return

        col_novo_nome = self.cb_col_novo_nome.currentText()
        if col_novo_nome == "(nenhuma)":
            col_novo_nome = ""

        col_pasta_destino = self.cb_col_pasta_destino.currentText()
        if col_pasta_destino == "(nenhuma)":
            col_pasta_destino = ""

        # Salva seleções de coluna
        self.cfg["col_matching"] = col_matching
        self.cfg["col_novo_nome"] = col_novo_nome
        self.cfg["col_pasta_destino"] = col_pasta_destino
        salvar_config(self.cfg)

        threshold = self.slider_threshold.value()
        sufixos = self.cfg.get("sufixos", CONFIG_PADRAO["sufixos"])

        # O v2 recebe a coluna ja como lista de strings — ele nao conhece pandas.
        self.motor = MotorMatching(
            coluna_como_texto(self.df, col_matching),
            sufixos,
            PesosScore(),
        )

        self.barra_varredura.setVisible(True)
        self.barra_varredura.setValue(0)
        self.status.showMessage("Escaneando arquivos…")

        self.thread_varredura = ThreadVarredura(
            pasta_raiz, self.motor, self.df, sufixos, threshold,
            col_novo_nome, col_pasta_destino
        )
        self.thread_varredura.progresso.connect(self.barra_varredura.setValue)
        self.thread_varredura.resultado_pronto.connect(self._on_resultado_pronto)
        self.thread_varredura.sem_match.connect(self.aba_sem_match.popular)
        self.thread_varredura.erro.connect(self._on_erro_varredura)
        self.thread_varredura.start()

    def _on_resultado_pronto(self, correspondencias: list[dict]):
        self.barra_varredura.setVisible(False)
        self.aba_correspondencias.popular_tabela(correspondencias)
        total_sem = len([
            f for f in self.aba_sem_match.lista.toPlainText().split("\n") if f.strip()
        ])
        self.abas.setCurrentIndex(0)
        self.status.showMessage(
            f"Varredura concluída: {len(correspondencias)} correspondências encontradas."
        )

    def _on_erro_varredura(self, msg: str):
        self.barra_varredura.setVisible(False)
        QMessageBox.critical(self, "Erro na varredura", msg)

    def _executar_acao(self, acao: str, pasta_base_destino: str):
        selecionados = self.aba_correspondencias.obter_selecionados()
        if not selecionados:
            QMessageBox.information(self, "Nenhum item", "Nenhum item selecionado.")
            return

        # Confirma com preview
        dialogo = DialogConfirmacao(selecionados, acao, pasta_base_destino, self)
        if dialogo.exec() != QDialog.Accepted:
            return

        self.aba_log.marcar_inicio_lote()
        self.abas.setCurrentIndex(3)  # Vai para aba Log

        prog = self.aba_correspondencias.progresso
        prog.setVisible(True)
        prog.setValue(0)

        self.thread_acao = ThreadAcao(selecionados, acao, pasta_base_destino)
        self.thread_acao.progresso.connect(prog.setValue)
        self.thread_acao.log_entrada.connect(self.aba_log.adicionar_entrada)
        self.thread_acao.concluido.connect(self._on_acao_concluida)
        self.thread_acao.start()

    def _on_acao_concluida(self, sucessos: int, erros: int):
        self.aba_correspondencias.progresso.setVisible(False)
        self.status.showMessage(
            f"Operação concluída: {sucessos} arquivo(s) com sucesso, {erros} erro(s)."
        )
        QMessageBox.information(
            self, "Concluído",
            f"✔ {sucessos} arquivo(s) processado(s) com sucesso.\n"
            f"{'⚠ ' + str(erros) + ' erro(s). Verifique o Log.' if erros else ''}"
        )

    def _on_config_alterada(self, nova_cfg: dict):
        self.cfg = nova_cfg

    def closeEvent(self, event):
        salvar_config(self.cfg)
        event.accept()


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def main():
    app = QApplication(sys.argv)
    app.setApplicationName("Organizador de Arquivos")
    app.setStyle("Fusion")

    janela = JanelaPrincipal()
    janela.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()


# ─────────────────────────────────────────────────────────────────────────────
# INSTRUÇÕES DE INSTALAÇÃO E BUILD
# ─────────────────────────────────────────────────────────────────────────────
#
# 1. Instalar dependências:
#    pip install PySide6 pandas openpyxl rapidfuzz chardet pyinstaller
#
# 2. Gerar o executável .exe (sem janela de console):
#    pyinstaller --onefile --windowed --name "OrganizadorArquivos" main.py
#
# 3. O .exe será gerado em:  dist/OrganizadorArquivos.exe
#    O config.json será criado na mesma pasta do .exe na primeira execução.
#
# ─────────────────────────────────────────────────────────────────────────────
